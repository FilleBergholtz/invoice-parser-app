"""Unit tests for Excel export."""

import pytest
import openpyxl
from pathlib import Path
from tempfile import TemporaryDirectory

from src.models.invoice_line import InvoiceLine
from src.models.row import Row
from src.models.segment import Segment
from src.export.excel_export import export_to_excel


@pytest.fixture
def sample_invoice_lines():
    """Create sample InvoiceLine objects for testing."""
    from src.models.page import Page
    from src.models.document import Document
    from src.models.token import Token
    
    doc = Document(filename="test.pdf", filepath="test.pdf", page_count=0, pages=[])
    page = Page(page_number=1, width=612, height=792, document=doc)
    doc.pages.append(page)
    doc.page_count = 1
    
    # Create rows with tokens first
    token1 = Token(text="Product", x=0, y=100, width=60, height=12, page=page)
    token2 = Token(text="Service", x=0, y=120, width=60, height=12, page=page)
    
    row1 = Row(tokens=[token1], text="Product 1", x_min=0, x_max=200, y=100, page=page)
    row2 = Row(tokens=[token2], text="Service 1", x_min=0, x_max=200, y=120, page=page)
    
    segment = Segment(
        segment_type="items",
        rows=[row1, row2],
        page=page,
        y_min=237.6,
        y_max=554.4
    )
    
    return [
        InvoiceLine(rows=[row1], description="Product 1", total_amount=60.0, line_number=1, segment=segment),
        InvoiceLine(rows=[row2], description="Service 1", total_amount=40.0, line_number=2, segment=segment),
    ]


def test_excel_export_with_control_columns(sample_invoice_lines, tmp_path):
    """Test Excel export includes control columns."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        "status": "OK",
        "lines_sum": 100.0,
        "diff": 0.0,
        "invoice_number_confidence": 0.96,
        "total_confidence": 0.97,
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    # Verify file was created
    assert output_file.exists()
    
    # Read Excel file
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Verify control columns exist
    headers = [cell.value for cell in ws[1]]
    assert "Status" in headers
    assert "Radsumma" in headers
    assert "Avvikelse" in headers
    assert "Fakturanummer-konfidens" in headers
    assert "Totalsumma-konfidens" in headers
    
    # Verify column order: existing + Fakturatotal, then control columns
    existing_cols = ["Fakturanummer", "Referenser", "Företag", "Fakturadatum",
                     "Beskrivning", "Antal", "Enhet", "Á-pris", "Rabatt", "Summa", "Hela summan", "Fakturatotal"]
    control_cols = ["Status", "Radsumma", "Avvikelse", "Fakturanummer-konfidens", "Totalsumma-konfidens"]

    assert headers[:len(existing_cols)] == existing_cols
    assert headers[len(existing_cols):] == control_cols


def test_control_column_values(sample_invoice_lines, tmp_path):
    """Test control column values match metadata."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        "status": "OK",
        "lines_sum": 100.0,
        "diff": 0.0,
        "invoice_number_confidence": 0.96,
        "total_confidence": 0.97,
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Check first data row (row 2, since row 1 is headers)
    row_data = [cell.value for cell in ws[2]]
    
    # Verify control column values
    # Column order: ... Hela summan (10), Fakturatotal (11), Status (12), Radsumma (13), Avvikelse (14), ...
    assert row_data[12] == "OK"  # Status
    assert row_data[13] == 100.0  # Radsumma
    assert row_data[14] == 0.0  # Avvikelse
    assert row_data[15] == 0.96  # Fakturanummer-konfidens
    assert row_data[16] == 0.97  # Totalsumma-konfidens


def test_control_columns_formatting(sample_invoice_lines, tmp_path):
    """Test control columns formatting (percentage for confidence, currency for amounts)."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        "status": "OK",
        "lines_sum": 100.0,
        "diff": 0.0,
        "invoice_number_confidence": 0.96,
        "total_confidence": 0.97,
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Check formatting for data rows
    for row_idx in range(2, ws.max_row + 1):
        row = ws[row_idx]
        
        # Radsumma at index 13 (after Fakturatotal 11, Status 12)
        radsumma_cell = row[13]
        assert radsumma_cell.number_format == "0.00"

        # Avvikelse at index 14
        avvikelse_cell = row[14]
        if isinstance(avvikelse_cell.value, (int, float)):
            assert avvikelse_cell.number_format == "0.00"

        # Confidence columns at 15, 16
        fakturanummer_konfidens_cell = row[15]
        assert fakturanummer_konfidens_cell.number_format == "0.00%"

        totalsumma_konfidens_cell = row[16]
        assert totalsumma_konfidens_cell.number_format == "0.00%"


def test_diff_n_a_handling(sample_invoice_lines, tmp_path):
    """Test that diff column shows 'N/A' when total_amount is None."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        "status": "REVIEW",
        "lines_sum": 100.0,
        "diff": "N/A",  # When total_amount is None
        "invoice_number_confidence": 0.90,
        "total_confidence": 0.85,
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Check that diff shows "N/A"
    row_data = [cell.value for cell in ws[2]]
    # Column order: ... Hela summan (10), Fakturatotal (11), Status (12), Radsumma (13), Avvikelse (14)
    assert row_data[14] == "N/A"  # Avvikelse

    # Verify "N/A" cell value is correct
    avvikelse_cell = ws[2][14]  # Avvikelse
    assert avvikelse_cell.value == "N/A"
    # Cell should contain text "N/A", not a number
    assert isinstance(avvikelse_cell.value, str)


def test_backward_compatibility(sample_invoice_lines, tmp_path):
    """Test that export works without validation fields (backward compatible)."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        # No validation fields
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Verify control columns use defaults (indices 12–16 after Fakturatotal 11)
    row_data = [cell.value for cell in ws[2]]
    assert row_data[12] == "REVIEW"  # Status default
    assert row_data[13] == 0.0  # lines_sum default
    assert row_data[14] == "N/A"  # diff default
    assert row_data[15] == 0.0  # invoice_number_confidence default
    assert row_data[16] == 0.0  # total_confidence default


def test_control_columns_repeat_per_invoice(sample_invoice_lines, tmp_path):
    """Test that control columns repeat same value for all rows of same invoice."""
    output_file = tmp_path / "test_invoices.xlsx"
    
    metadata = {
        "fakturanummer": "INV-001",
        "foretag": "Test Company",
        "fakturadatum": "2024-01-15",
        "status": "OK",
        "lines_sum": 100.0,
        "diff": 0.0,
        "invoice_number_confidence": 0.96,
        "total_confidence": 0.97,
    }
    
    export_to_excel(sample_invoice_lines, str(output_file), metadata)
    
    wb = openpyxl.load_workbook(output_file)
    ws = wb['Invoices']
    
    # Verify all rows have same control column values (2 invoice lines); indices 12–16
    for row_idx in range(2, ws.max_row + 1):
        row_data = [cell.value for cell in ws[row_idx]]

        assert row_data[12] == "OK"  # Status
        assert row_data[13] == 100.0  # Radsumma
        assert row_data[14] == 0.0  # Avvikelse
        assert row_data[15] == 0.96  # Fakturanummer-konfidens
        assert row_data[16] == 0.97  # Totalsumma-konfidens
